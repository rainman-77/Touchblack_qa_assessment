import os
import openpyxl
from configurations.read_configurations import read_configuration as rc


def load_workbook_and_sheet(sheet_name):
    excel_path = rc("excel info", "path")
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]
    return workbook, sheet

def get_row_count(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.max_row


def get_column_count(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.max_column


def get_cell_data(sheet_name, row_num, column_num):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.cell(row_num, column_num).value


def set_cell_data(sheet_name, row_num, column_num, data):
    excel_path = rc("excel info", "path")
    workbook, sheet = load_workbook_and_sheet(sheet_name)
    sheet.cell(row_num, column_num).value = data
    workbook.save(excel_path)


def get_all_excel_data(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    final_list = []
    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)
    return final_list       # list of lists is the output
